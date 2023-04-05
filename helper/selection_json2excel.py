#

import json
import openpyxl
import re

################################################################
# INPUT
YEAR = 2023
filename = "../CERT-Officer-Dienstplan_{}.xlsx".format(YEAR)
################################################################

sheetname_regex = '{}-0[1-9]|1[012]'.format(YEAR)

if __name__ == '__main__':
    excel_workbook = openpyxl.load_workbook(filename=filename)

    officer_selection = json.load(open("../data/officer_selection.json"))
    if not officer_selection:
        raise Exception('There is no officer selection data')

    # noinspection SpellCheckingInspection
    for sheetname in excel_workbook.sheetnames:
        # noinspection DuplicatedCode
        if not re.match(sheetname_regex, sheetname):
            continue

        sheet = excel_workbook[sheetname]

        if not sheet.cell(row=1, column=1).value == YEAR:
            print(sheet.cell(row=1, column=1).value)
            raise Exception

        officers = {}
        lastname = 'notNone'
        current_row = 3
        while lastname:
            lastname = sheet.cell(row=current_row, column=1).value
            if lastname:
                officers[lastname] = current_row
            current_row += 1

        for officer, row in officers.items():
            for column in range(1, 31 + 1):
                if not sheet.cell(row=1, column=column + 2).value == column:
                    continue

                if officer_selection.get('{}-{:02}'.format(sheetname, column)):
                    if officer.strip() in officer_selection.get('{}-{:02}'.format(sheetname, column)):
                        sheet.cell(row=row, column=column + 2).value = 'x'

    excel_workbook.save(filename)
