#

import json
import re
from datetime import datetime

import openpyxl

################################################################
# INPUT
YEAR = 2023
source_filename = "../CERT-Officer-Dienstplan_{}.xlsx".format(YEAR)
destination_filename = "../CERT-Officer-Dienstplan_Auswertung_{}.xlsx".format(YEAR)
blocked = []
################################################################
sheetname_regex = '({})-(0[1-9]|1[012])'.format(YEAR)

if __name__ == '__main__':
    bridging_days = json.load(open("../data/bridging_days.json"))  # source for bridging_days
    if not bridging_days:
        raise Exception('There is no bridging days data')

    officer_dynamics_general_normal = {}
    officer_dynamics_general_extreme = {}
    officer_dynamics_monthly_normal = {}
    officer_dynamics_monthly_extreme = {}

    excel_source_workbook = openpyxl.load_workbook(filename=source_filename)

    # try:
    #     excel_destination_workbook = openpyxl.load_workbook(filename=destination_filename)
    # except FileNotFoundError:
    #     excel_destination_workbook = openpyxl.Workbook()
    #     # excel_destination_workbook.remove_sheet('Sheet')
    excel_destination_workbook = openpyxl.Workbook()

    try:
        excel_worksheet_general = excel_destination_workbook["Officer Times"]
    except KeyError:
        excel_worksheet_general = excel_destination_workbook.create_sheet("Officer Times")

    try:
        excel_worksheet_normal = excel_destination_workbook["Officer Times Monthy"]
    except KeyError:
        excel_worksheet_normal = excel_destination_workbook.create_sheet("Officer Times Monthy")

    try:
        excel_worksheet_extreme = excel_destination_workbook["Officer Extreme Times Monthly"]
    except KeyError:
        excel_worksheet_extreme = excel_destination_workbook.create_sheet("Officer Extreme Times Monthly")

    excel_worksheet_general.cell(row=2, column=1).value = "Normal"
    excel_worksheet_general.cell(row=3, column=1).value = "Extreme"

    for sheetname in excel_source_workbook.sheetnames:
        match = re.fullmatch(sheetname_regex, sheetname)
        if not match:
            continue
        groups = match.groups()
        year = groups[0]
        month = groups[1]

        sheet = excel_source_workbook[sheetname]

        if not sheet.cell(row=1, column=1).value == YEAR:
            print(sheet.cell(row=1, column=1).value)
            raise Exception

        officers = {}
        lastname = 'notNone'
        current_row = 3
        while lastname:
            lastname = sheet.cell(row=current_row, column=1).value
            if lastname:
                officers[lastname.strip()] = current_row
            current_row += 1

        officers = {officer: row for officer, row in officers.items() if officer not in blocked}

        if not officer_dynamics_monthly_normal.get(month):
            officer_dynamics_monthly_normal[month] = {}

        if not officer_dynamics_monthly_extreme.get(month):
            officer_dynamics_monthly_extreme[month] = {}

        for officer, row in officers.items():
            if not officer_dynamics_general_normal.get(officer.strip()):
                officer_dynamics_general_normal[officer.strip()] = 0

            if not officer_dynamics_general_extreme.get(officer.strip()):
                officer_dynamics_general_extreme[officer.strip()] = 0

            if not officer_dynamics_monthly_normal[month].get(officer.strip()):
                officer_dynamics_monthly_normal[month][officer.strip()] = 0

            if not officer_dynamics_monthly_extreme[month].get(officer.strip()):
                officer_dynamics_monthly_extreme[month][officer.strip()] = 0

            for column in range(1, 31 + 1):
                if sheet.cell(row=1, column=column + 2).value != column:
                    continue

                if sheet.cell(row=row, column=column + 2).value != 'x':
                    continue

                day_str = '{}-{:02}'.format(sheetname, column)
                day = datetime.strptime(day_str, "%Y-%m-%d")
                month = sheetname.split('-')[1]

                officer_dynamics_general_normal[officer] += 1
                officer_dynamics_monthly_normal[month][officer] += 1

                if day.weekday() in [0, 4] or day.strftime("YYYY-MM-DD") in bridging_days:
                    officer_dynamics_general_extreme[officer] += 1
                    officer_dynamics_monthly_extreme[month][officer] += 1

    excel_worksheet_general.cell(row=1, column=2).value = "Normal"
    for i, (officer, count) in enumerate(officer_dynamics_general_normal.items()):
        excel_worksheet_general.cell(row=i + 2, column=1).value = officer
        excel_worksheet_general.cell(row=i + 2, column=2).value = count

    excel_worksheet_general.cell(row=1, column=3).value = "Extreme"
    for i, (officer, count) in enumerate(officer_dynamics_general_extreme.items()):
        excel_worksheet_general.cell(row=i + 2, column=3).value = count

    officer_row_map_normal = {}
    for i, (month, officers) in enumerate(officer_dynamics_monthly_normal.items()):
        excel_worksheet_normal.cell(row=1, column=i + 2).value = month
        for y, (officer, count) in enumerate(officers.items()):
            if i == 0:
                officer_row_map_normal[officer] = y + 2
                excel_worksheet_normal.cell(row=officer_row_map_normal.get(officer), column=1).value = officer
            if officer_row_map_normal.get(officer):
                excel_worksheet_normal.cell(row=officer_row_map_normal.get(officer), column=i + 2).value = count
            else:
                officer_row_map_normal[officer] = len(officer_row_map_normal) + 2
                excel_worksheet_normal.cell(row=officer_row_map_normal.get(officer), column=1).value = officer
                excel_worksheet_normal.cell(row=officer_row_map_normal.get(officer), column=i + 2).value = count

    officer_row_map_extreme = {}
    for i, (month, officers) in enumerate(officer_dynamics_monthly_extreme.items()):
        excel_worksheet_extreme.cell(row=1, column=i + 2).value = month
        for y, (officer, count) in enumerate(officers.items()):
            if i == 0:
                officer_row_map_extreme[officer] = y + 2
                excel_worksheet_extreme.cell(row=officer_row_map_extreme.get(officer), column=1).value = officer
            if officer_row_map_extreme.get(officer):
                excel_worksheet_extreme.cell(row=officer_row_map_extreme.get(officer), column=i + 2).value = count
            else:
                officer_row_map_extreme[officer] = len(officer_row_map_extreme) + 2
                excel_worksheet_extreme.cell(row=officer_row_map_extreme.get(officer), column=1).value = officer
                excel_worksheet_extreme.cell(row=officer_row_map_extreme.get(officer), column=i + 2).value = count

    excel_destination_workbook.remove(excel_destination_workbook['Sheet'])
    excel_destination_workbook.save(destination_filename)
