#

import datetime
import json
import openpyxl
from openpyxl.styles import PatternFill

################################################################
# INPUT
year = 2023
month = 12
month_name = 'Dezember'
filename = "../CERT-Officer-Dienstplan_{}.xlsx".format(year)
################################################################

sheetname = f'{year}-{month:02}'
day_map = ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So']
holidayColor = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

if __name__ == '__main__':

    excel_workbook = openpyxl.load_workbook(filename=filename)

    template_excel_worksheet = excel_workbook["TEMPLATE"]

    officer_static = json.load(open("../data/officer_static.json"))
    if not officer_static:
        raise Exception('There is no officer static data')

    sheet = excel_workbook.copy_worksheet(template_excel_worksheet)

    sheet.title = sheetname

    sheet.cell(row=1, column=1).value = year
    sheet.cell(row=1, column=2).value = month_name

    for i, officer in enumerate(officer_static):
        sheet.cell(row=i+3, column=1).value = officer.get('name')

    if month == 12:
        latest_day_in_month = datetime.datetime(year=year+1, month=1, day=1) - datetime.timedelta(days=1)
    else:
        latest_day_in_month = datetime.datetime(year=year, month=month + 1, day=1) - datetime.timedelta(days=1)

    for i in range(1, latest_day_in_month.day+1):
        weekday = datetime.datetime(year=year, month=month, day=i).weekday()

        sheet.cell(row=1, column=i+2).value = i
        sheet.cell(row=2, column=i+2).value = day_map[weekday]

        if weekday in [5, 6]:
            for y in range(1, len(officer_static)+3):
                sheet.cell(row=y, column=i+2).fill = holidayColor

    for i in range(latest_day_in_month.day, 31+1):
        sheet.cell(row=1, column=i+3).value = ""

    excel_workbook.save(filename)
