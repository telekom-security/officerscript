#

import json
import re
from datetime import datetime

import openpyxl

from model import Officer

################################################################
# INPUT
YEAR = 2023
FIRST_MONTH = 7
LAST_MONTH = 7
filename = "../CERT-Officer-Dienstplan_{}.xlsx".format(YEAR)
################################################################

sheetname_regex = '({})-(0[1-9]|1[012])'.format(YEAR)


def get_officer_by_name(name, _officers):
    for _officer in _officers:
        if _officer.name == name:
            return _officer


if __name__ == '__main__':
    officer_selection = {}
    officer_dynamics = {}
    excel_workbook = openpyxl.load_workbook(filename=filename)

    officer_static_data = json.load(open("../data/officer_static.json"))  # source for officer static data
    if not officer_static_data:
        raise Exception('There is no officer static data')

    officer_objs = []
    for officer in officer_static_data:
        officer_objs.append(Officer(officer))

    bridging_days = json.load(open("../data/bridging_days.json"))  # source for bridging_days
    if not bridging_days:
        raise Exception('There is no bridging days data')

    # noinspection SpellCheckingInspection
    for sheetname in excel_workbook.sheetnames:
        match = re.fullmatch(sheetname_regex, sheetname)
        if not match:
            continue
        groups = match.groups()
        year = groups[0]
        month = groups[1]

        if not int(month) >= FIRST_MONTH:
            continue

        if not int(month) <= LAST_MONTH:
            continue

        sheet = excel_workbook[sheetname]

        if not sheet.cell(row=1, column=1).value == YEAR:
            print(sheet.cell(row=1, column=1).value)
            raise Exception

        officers = {}
        lastname = 'notNone'
        current_row = 3
        while lastname:
            lastname = sheet.cell(row=current_row, column=1).value
            if lastname:
                officers[lastname.strip()] = current_row
            current_row += 1

        for officer_name, row in officers.items():
            officer_obj = get_officer_by_name(officer_name, officer_objs)

            if not officer_obj:
                continue

            if not officer_dynamics.get(officer_name):
                officer_dynamics[officer_name] = {'officer_count': 0, 'officer_extreme_count': 0}

            for column in range(1, 31 + 1):
                if sheet.cell(row=1, column=column + 2).value != column:
                    continue

                if sheet.cell(row=row, column=column + 2).value != 'x':
                    continue

                day_str = '{}-{:02}'.format(sheetname, column)
                day = datetime.strptime(day_str, "%Y-%m-%d")

                if day_str not in officer_selection:
                    officer_selection[day_str] = []

                officer_dynamics[officer_name]['officer_count'] += 1

                if day.weekday() in [0, 4] or day.strftime("YYYY-MM-DD") in \
                        bridging_days.get(officer_obj.country).get(officer_obj.state):
                    officer_dynamics[officer_name]['officer_extreme_count'] += 1

                officer_selection[day_str].append(officer_name)

    # noinspection PyTypeChecker
    officer_selection = dict(sorted(officer_selection.items()))

    with open('../data/officer_selection.json', 'w') as f:
        f.write(json.dumps(officer_selection, indent=4))

    with open('../data/officer_dynamic.json', 'w') as f:
        f.write(json.dumps(officer_dynamics, indent=4))
