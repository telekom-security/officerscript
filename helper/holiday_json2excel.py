#

import datetime
import json
import openpyxl
import re
from openpyxl.styles import PatternFill


################################################################
# INPUT
YEAR = 2023
filename = "../CERT-Officer-Dienstplan_{}.xlsx".format(YEAR)
################################################################

sheetname_regex = '{year}-0[1-9]|{year}-1[012]'.format(year=YEAR)

holidayColor = PatternFill(start_color='F8CBAD',
                           end_color='F8CBAD',
                           fill_type='solid')

if __name__ == '__main__':
    excel_workbook = openpyxl.load_workbook(filename=filename)

    holidays = json.load(open("../data/holidays.json"))
    if not holidays:
        raise Exception('There is no holidays data')

    officer_static_data = json.load(open("../data/officer_static.json"))  # source for officer static data
    if not officer_static_data:
        raise Exception('There is no officer static data')

    officers_with_data = {}
    for officer in officer_static_data:
        officers_with_data[officer.get('name')] = officer

    # noinspection SpellCheckingInspection
    for sheetname in excel_workbook.sheetnames:
        # noinspection DuplicatedCode
        if not re.match(sheetname_regex, sheetname):
            continue

        sheet = excel_workbook[sheetname]

        if not sheet.cell(row=1, column=1).value == YEAR:
            print(sheet.cell(row=1, column=1).value)
            raise Exception("Something went really wrong")

        officers = {}
        lastname = 'notNone'
        current_row = 3
        while lastname:
            lastname = sheet.cell(row=current_row, column=1).value
            if lastname:
                officers[lastname.strip()] = current_row
            current_row += 1

        for officer, row in officers.items():
            if not officers_with_data.get(officer):
                continue

            country = officers_with_data.get(officer).get('country')
            state = officers_with_data.get(officer).get('state')

            _holidays = holidays.get(country).get(state)

            for column in range(1, 31 + 1):
                if not sheet.cell(row=1, column=column + 2).value == column:
                    continue

                if datetime.datetime.strptime('{}-{:02}'.format(sheetname, column), '%Y-%m-%d').weekday() in [5, 6]:
                    sheet.cell(row=row, column=column + 2).fill = holidayColor

                if '{}-{:02}'.format(sheetname, column) in _holidays:
                    sheet.cell(row=row, column=column + 2).fill = holidayColor

    excel_workbook.save(filename)
